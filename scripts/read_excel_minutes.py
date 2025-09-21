#!/usr/bin/env python3
"""
Excel（.xlsx）の議事録をMarkdownに変換するユーティリティ。

特徴
- サブフォルダも含めて .xlsx を再帰検索（~$ で始まる一時ファイルは除外）
- シートごとにMarkdownの章（## 見出し）を作成
- 1ファイルに集約 or ファイルごとに分割の両対応
- 依存は openpyxl のみ（pandas不要）

使い方（例）
  python scripts/read_excel_minutes.py --base-dir "C:\\Users\\saway\\okinakajyuku\\okinakajyuku\\25年度議事録" \
    --output minutes_summary.md

  python scripts/read_excel_minutes.py --base-dir "D:\\share\\25年度議事録" \
    --per-file --output-dir docs/minutes

事前に: pip install openpyxl
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import List, Optional

from openpyxl import load_workbook


def escape_md(text: Optional[object]) -> str:
    """Markdown表向けにセル値を安全化。Noneは空文字。改行は<br>、'|'はエスケープ。"""
    if text is None:
        return ""
    s = str(text)
    s = s.replace("|", r"\|")
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\n", "<br>")
    return s


def first_nonempty_row(rows: List[List[object]]) -> int:
    for idx, row in enumerate(rows):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            return idx
    return 0


def trim_trailing_empty_rows(rows: List[List[object]]) -> List[List[object]]:
    i = len(rows) - 1
    while i >= 0:
        if any(cell is not None and str(cell).strip() != "" for cell in rows[i]):
            break
        i -= 1
    return rows[: i + 1]


def sheet_to_markdown(ws) -> Optional[str]:
    """ワークシートをMarkdown表に変換。空ならNone。"""
    # openpyxlは .values で行イテレータを返す
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    rows = trim_trailing_empty_rows(rows)
    if not rows:
        return None

    start = first_nonempty_row(rows)
    rows = rows[start:]
    if not rows:
        return None

    header = [escape_md(c) for c in rows[0]]
    data = [[escape_md(c) for c in r] for r in rows[1:]]

    # 列数を最大行に合わせて揃える
    max_cols = max(len(header), *(len(r) for r in data) if data else [len(header)])
    if len(header) < max_cols:
        header += [""] * (max_cols - len(header))
    for r in data:
        if len(r) < max_cols:
            r += [""] * (max_cols - len(r))

    # Markdown表の構築
    parts = []
    parts.append("| " + " | ".join(header) + " |")
    parts.append("| " + " | ".join(["---"] * max_cols) + " |")
    for r in data:
        parts.append("| " + " | ".join(r) + " |")
    return "\n".join(parts)


def extract_highlights(header: List[str], data: List[List[str]]):
    """決定事項/宿題/ToDo 風の列があれば箇条書きにして返す。なければ空。"""
    lowers = [h.lower() for h in header]
    picks = {
        "decisions": [i for i, h in enumerate(lowers) if "決定" in h or "decision" in h],
        "todos": [i for i, h in enumerate(lowers) if "宿題" in h or "todo" in h],
        "notes": [i for i, h in enumerate(lowers) if "メモ" in h or "note" in h],
    }
    out = []
    for label, idxs in picks.items():
        buf = []
        for row in data:
            for i in idxs:
                if i < len(row):
                    v = row[i]
                    if v and str(v).strip():
                        buf.append(f"- {escape_md(v)}")
        if buf:
            title = {"decisions": "決定事項", "todos": "宿題/ToDo", "notes": "メモ"}[label]
            out.append(f"### {title}\n" + "\n".join(buf))
    return "\n\n".join(out)


def process_file(xlsx_path: Path) -> List[str]:
    parts: List[str] = [f"# {xlsx_path.name}"]
    try:
        wb = load_workbook(filename=xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        parts.append(f"> 読み込みエラー: {e}")
        return parts

    for ws in wb.worksheets:
        md = sheet_to_markdown(ws)
        if md:
            parts.append(f"## {ws.title}")
            parts.append(md)
    return parts


def render_single(base_dir: Path, output: Path) -> None:
    files = [p for p in sorted(base_dir.rglob("*.xlsx")) if not p.name.startswith("~$")]
    lines: List[str] = [f"# 議事録まとめ（{base_dir}）", ""]
    for x in files:
        lines.extend(process_file(x))
        lines.append("")
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text("\n".join(lines), encoding="utf-8")
    print(f"Wrote {output}")


def render_per_file(base_dir: Path, out_dir: Path) -> None:
    files = [p for p in sorted(base_dir.rglob("*.xlsx")) if not p.name.startswith("~$")]
    out_dir.mkdir(parents=True, exist_ok=True)
    for x in files:
        lines = process_file(x)
        dest = out_dir / f"{x.stem}.md"
        dest.write_text("\n".join(lines), encoding="utf-8")
        print(f"Wrote {dest}")


def main():
    ap = argparse.ArgumentParser(description="Excel議事録（.xlsx）をMarkdownへ変換")
    ap.add_argument("--base-dir", required=True, help=".xlsx が入っているフォルダ（再帰検索）")
    ap.add_argument("--output", default="minutes_summary.md", help="集約出力のファイル名（--per-file無し時）")
    ap.add_argument("--per-file", action="store_true", help="ファイルごとに個別Markdownを出力")
    ap.add_argument("--output-dir", default="minutes_output", help="--per-file時の出力先フォルダ")
    args = ap.parse_args()

    base = Path(args.base_dir)
    if not base.exists():
        raise SystemExit(f"base-dir が見つかりません: {base}")

    if args.per_file:
        render_per_file(base, Path(args.output_dir))
    else:
        render_single(base, Path(args.output))


if __name__ == "__main__":
    main()

